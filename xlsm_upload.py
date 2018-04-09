import openpyxl
import sys
import json

# book = openpyxl.load_workbook('C:\\hosting\\mdglc\\vice_dev\\python\\test_eqps.xlsm')
# print(str(sys.argv))

# print("xlsm_upload.py")

book = openpyxl.load_workbook(sys.argv[1])
sheet = book.get_sheet_by_name('Sheet1')

result_array = {}
count = 0
for i, row in enumerate(sheet.iter_rows()):
    result_row = {}
    if (i == 0):
        keys = {}
        for j, c in enumerate(row):
            keys[j] = c.value
#            print(str(j) + ':  '+keys[j])
    if (i > 0):
        values = {}
        for j, c in enumerate(row):
            values[j] = c.value
#            print(keys[j]+':  '+str(values[j]))
            result_row[keys[j]] = str(values[j])
        result_array[count] = result_row
        count = count + 1
#    print('\n')

result_json = json.dumps(result_array)
print(result_json)